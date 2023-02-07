from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title='HCL'
#sheet['A1'].value=10
#sheet['B2'].value='test'
#sheet.cell(row=3,column=3).value='HCL DATA'
j=0
for i in range(10,60,10):
    j+=1
    sheet.cell(row=j,column=1).value=i

for row in sheet.iter_rows(min_row=1,max_row=5,max_col=2,min_col=2):
    for r in row:
        r.value=1

i=0
for column in sheet.iter_cols(min_row=6,max_row=6,max_col=5,min_col=1):
    for r in column:
        i+=100
        r.value=i
wb.save("C://Users/user786/Desktop/hcl.xlsx")