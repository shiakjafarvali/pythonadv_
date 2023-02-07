import openpyxl
import openpyxl as opx
wb=openpyxl.load_workbook("C://Users/user786/Desktop/Book2.xlsx")
sheet=wb.active
print(sheet)
data=sheet['A3'].value
data3=sheet.cell(row=2,column=3).value
data1=sheet['A1:A10']
print(data)
print(data1)
print(data3)
#access all cells in row1
print(sheet.max_row)
print(sheet.max_column)
for row in sheet.rows:
    print([data.value for data in row])
for i in range(1,12):
    print(sheet.cell(row=i,column=1).value)
for i in range(2,12):
    print(sheet.cell(row=i,column=1).value)
sheet["h2"]='SUM(C2:G2)'